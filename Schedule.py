from Student import Student
import openpyxl


class Schedule:
    __path = 'schedule.xlsx'
    __wb_obj = openpyxl.load_workbook(__path)
    __sheet_obj = __wb_obj.active
    __row = __sheet_obj.max_row
    __column = __sheet_obj.max_column

    def GetShedule(self, id, week):
        daysOfWeek = []
        schedule = []
        endEvenWeek = self.__row // 2 + 1
        startOddWeek = self.__row // 2 + 3
        for i in range(1, self.__column + 1):
            days = self.__sheet_obj.cell(row=1, column=i)
            daysOfWeek += [int(days.value)]
        if week == 'Чётная':
            for j in range(1, self.__column + 1):
                lessons = []
                for i in range(2, endEvenWeek):
                    cell_obj = self.__sheet_obj.cell(row=i, column=j)
                    if cell_obj.value != None:
                        lessons += [cell_obj.value]
                    else:
                        lessons += ['---']

                schedule += [{daysOfWeek[j - 1]: lessons}]
            return schedule
        if week == 'Нечётная':
            for j in range(1, self.__column + 1):
                lessons = []
                for i in range(startOddWeek, self.__row + 1):
                    cell_obj = self.__sheet_obj.cell(row=i, column=j)
                    if cell_obj.value != None:
                        lessons += [cell_obj.value]
                    else:
                        lessons += ['---']

                schedule += [{daysOfWeek[j - 1]: lessons}]
            return schedule
        else:
            raise ValueError('Некорректная неделя!')


    def GetDayOfSchedule(self,day,numberWeek):
        Day = self.GetShedule(id,numberWeek)
        return Day[day-1]

    def PrintSchedule(self, id, week, day):
        week = self.GetShedule(id, week)
        return week[day - 1][day]